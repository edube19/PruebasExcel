from openpyxl import load_workbook
import os
import random
def leer_libro(wb):
    #wb = load_workbook(ruta)
    #wb = load_workbook('BCP SOLES.xlsx')
    print(wb.sheetnames)#imprime las hojas q tiene el libro
 
# Print value of cell object
# using the value attribute
def imprimir_valores(fila,columna,wb):
    sheet_obj = wb.active
    print('leyendo la fila 5, columna 2')
    cell_obj = sheet_obj.cell(row = fila, column = columna)
    print(cell_obj.value)#imprime ese valor de la celda x=5, y=2
    print(sheet_obj.max_row)#imprime el total de filas
    print(sheet_obj.max_column)#imprime el total de columnas
    #El total lo define buscando si hay alguna valor en la ultima fila/columna
def imprimir_fila(fila,wb):
    sheet_obj = wb.active
    max_col = sheet_obj.max_column
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = fila, column = i)
        print(cell_obj.value)

def imprimir_columna(columna,wb):
    sheet_obj = wb.active
    m_row = sheet_obj.max_row
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = columna)
        print(cell_obj.value)

def editar_celda(celda,wb,ruta,nuevo_valor):
    wb = load_workbook(ruta)
    sheet=wb.active
    cell=sheet[celda]
    cell.value = sheet[f'{celda}'].value
    #sheet.cell(row=1, column=6).value = 2 otra forma
    print(cell.value)
    cell.value=nuevo_valor
    print(cell.value)
    wb.save(ruta)
#B11118:G11121
def editar_varias_celdas(celda_inicio,celda_final,wb,ruta):
    wb = load_workbook(ruta)
    sheet=wb.active
    for row in sheet[f'{celda_inicio}:{celda_final}']:
        for cell in row:
            numero = random.randint(1000,3250)
            cell.value=numero
            #print(cell.value)
    wb.save(ruta)

def leer_bloques_celdas(celda_inicio,celda_final,wb,ruta):
    wb = load_workbook(ruta)
    sheet=wb.active
    for row in sheet[f'{celda_inicio}:{celda_final}']:
        for cell in row:
            print(cell.value)
    wb.save(ruta)

"""
sheet=book.active
for row in sheet['b5:f5']:
    for cell in row:
        print(cell.value)
"""